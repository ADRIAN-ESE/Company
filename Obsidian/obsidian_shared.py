"""OBSIDIAN CORPORATION — Shared Data Layer v3.0"""
import json, os, csv, io, hashlib, secrets
from datetime import datetime

DATA_FILE  = "obsidian_data.json"
ROLE_LEVEL = {"admin": 4, "hr": 3, "manager": 2, "staff": 1}
ROLE_LABEL = {"admin": "Admin", "hr": "HR Manager", "manager": "Manager", "staff": "Staff"}

# ─── ROLE DEFINITIONS ────────────────────────────────────────────────────────
ROLE_DEFINITIONS = {
    "admin": {
        "label":       "System Administrator",
        "icon":        "🛡️",
        "color":       "#e05252",
        "purpose":     "Full system control and governance. Manages all users, configurations, and sensitive data.",
        "responsibilities": [
            "Create, edit, and delete system user accounts",
            "Configure access roles and permission levels",
            "Approve or override any HR decision",
            "Access all audit logs and system settings",
            "Manage department structure and org-wide policies",
        ],
        "level": 4,
    },
    "hr": {
        "label":       "HR Manager",
        "icon":        "👔",
        "color":       "#4f8ef7",
        "purpose":     "Manages the complete employee lifecycle from onboarding to offboarding and all people-operations workflows.",
        "responsibilities": [
            "Add, edit, and remove employee records",
            "Approve or deny promotion and leave requests",
            "View and manage payroll and salary data",
            "Run workforce analytics and export reports",
            "Coordinate training, benefits, and compliance",
        ],
        "level": 3,
    },
    "manager": {
        "label":       "Department Manager",
        "icon":        "📋",
        "color":       "#e09a30",
        "purpose":     "Oversees team performance and day-to-day operations within assigned departments.",
        "responsibilities": [
            "View all employee records and department data",
            "Submit and track promotion requests for team members",
            "Submit leave requests on behalf of direct reports",
            "Monitor team attendance, status, and workload",
            "Collaborate with HR on hiring and performance reviews",
        ],
        "level": 2,
    },
    "staff": {
        "label":       "Staff Member",
        "icon":        "🧑‍💼",
        "color":       "#3dba7a",
        "purpose":     "Standard employee portal access for self-service HR tasks and personal record management.",
        "responsibilities": [
            "View own employee profile and records",
            "Submit personal leave requests",
            "Request a promotion or role change",
            "View own payslip summary",
            "Update profile photo and change password",
        ],
        "level": 1,
    },
}

# ─── DEPARTMENT DEFINITIONS ───────────────────────────────────────────────────
DEPT_DEFINITIONS = {
    "Administration": {
        "icon":        "🏛️",
        "color":       "#c9a84c",
        "purpose":     "Serves as the operational backbone of the organisation, ensuring smooth corporate governance, executive support, and facility management.",
        "responsibilities": [
            "Corporate governance and board secretariat support",
            "Executive and C-suite administrative assistance",
            "Facilities, fleet, and office resource management",
            "Records, document control, and archiving",
            "Internal communications and correspondence routing",
        ],
        "kpis": ["Document turnaround time", "Facilities uptime", "Compliance adherence"],
    },
    "Human Resources": {
        "icon":        "👥",
        "color":       "#4f8ef7",
        "purpose":     "Drives talent strategy, employee engagement, and people operations to build and sustain a high-performing workforce.",
        "responsibilities": [
            "Talent acquisition, onboarding, and offboarding",
            "Performance management and career development",
            "Compensation, benefits, and payroll coordination",
            "Employee relations, welfare, and grievance handling",
            "Training, learning programmes, and policy compliance",
        ],
        "kpis": ["Time-to-hire", "Employee retention rate", "Training completion rate"],
    },
    "Finance": {
        "icon":        "💰",
        "color":       "#3dba7a",
        "purpose":     "Safeguards the financial health of the organisation through rigorous planning, reporting, and treasury management.",
        "responsibilities": [
            "Financial planning, budgeting, and forecasting",
            "Accounts payable, receivable, and reconciliation",
            "Internal and external audit coordination",
            "Treasury management and cash flow optimisation",
            "Regulatory reporting and statutory compliance",
        ],
        "kpis": ["Budget variance", "Audit findings count", "Cash flow coverage ratio"],
    },
    "Operations": {
        "icon":        "⚙️",
        "color":       "#e09a30",
        "purpose":     "Delivers core business services with operational excellence, ensuring quality, efficiency, and continuous improvement across all processes.",
        "responsibilities": [
            "End-to-end service delivery and process execution",
            "Logistics, supply chain, and vendor coordination",
            "Quality assurance and standards compliance",
            "Capacity planning and resource optimisation",
            "Process improvement and SOP development",
        ],
        "kpis": ["Service delivery rate", "Process cycle time", "Customer satisfaction score"],
    },
    "Technology": {
        "icon":        "💻",
        "color":       "#7c5df9",
        "purpose":     "Powers the organisation's digital infrastructure, software systems, and cybersecurity posture to enable business continuity and innovation.",
        "responsibilities": [
            "Software development, deployment, and maintenance",
            "IT infrastructure, cloud, and network management",
            "Cybersecurity, data protection, and incident response",
            "Data analytics, business intelligence, and reporting",
            "Digital transformation and technology strategy",
        ],
        "kpis": ["System uptime", "Incident resolution time", "Security vulnerability count"],
    },
}

def get_dept_definition(dept_name):
    if dept_name in DEPT_DEFINITIONS:
        return DEPT_DEFINITIONS[dept_name]
    return {
        "icon":  "🏢",
        "color": "#5a6080",
        "purpose": f"Responsible for {dept_name.lower()} activities and operations.",
        "responsibilities": ["Department-specific duties and responsibilities"],
        "kpis":  ["Performance metrics", "Output quality", "Team efficiency"],
    }

# ─── AUTH ─────────────────────────────────────────────────────────────────────
def hash_pw(pw):
    salt = secrets.token_hex(16)
    return f"{salt}:{hashlib.sha256((salt+pw).encode()).hexdigest()}"

def verify_pw(pw, stored):
    try:
        salt, h = stored.split(":", 1)
        return hashlib.sha256((salt+pw).encode()).hexdigest() == h
    except Exception:
        return False

# ─── DATA ─────────────────────────────────────────────────────────────────────
def _defaults():
    now = datetime.now().strftime("%Y-%m-%d")
    return {
        "meta": {"last_emp_num": 0},
        "system_users": {
            "admin":      {"password_hash": hash_pw("admin123"),  "role": "admin",   "full_name": "System Administrator", "emp_id": None, "photo": None, "created": now},
            "hr_manager": {"password_hash": hash_pw("hr123"),     "role": "hr",      "full_name": "HR Manager",           "emp_id": None, "photo": None, "created": now},
            "manager1":   {"password_hash": hash_pw("mgr123"),    "role": "manager", "full_name": "Department Manager",   "emp_id": None, "photo": None, "created": now},
            "staff1":     {"password_hash": hash_pw("staff123"),  "role": "staff",   "full_name": "Staff Member",         "emp_id": None, "photo": None, "created": now},
        },
        "employees": {},
        "departments": {
            "Administration":  {"description": "Administrative operations and corporate governance", "head_id": None},
            "Human Resources": {"description": "People operations, talent acquisition and development", "head_id": None},
            "Finance":         {"description": "Financial planning, accounting and treasury", "head_id": None},
            "Operations":      {"description": "Core business operations and service delivery", "head_id": None},
            "Technology":      {"description": "IT infrastructure, software and digital systems", "head_id": None},
        },
        "audit_log": [],
        "notifications": [],
        "leave_requests": [],
    }

def load_data():
    if not os.path.exists(DATA_FILE):
        d = _defaults(); save_data(d); return d
    try:
        with open(DATA_FILE) as f: return json.load(f)
    except Exception: return _defaults()

def save_data(data):
    with open(DATA_FILE, "w") as f: json.dump(data, f, indent=2)

def gen_emp_id(data):
    data["meta"]["last_emp_num"] = data["meta"].get("last_emp_num", 0) + 1
    return f"OBSD-{datetime.now().year}-{data['meta']['last_emp_num']:04d}"

def audit(data, user, action, target, detail=""):
    data.setdefault("audit_log", []).insert(0, {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": user, "action": action, "target": target, "detail": detail,
    })
    data["audit_log"] = data["audit_log"][:1000]

def notify(data, ntype, message, roles=None):
    data.setdefault("notifications", []).insert(0, {
        "id": secrets.token_hex(8), "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": ntype, "message": message, "read_by": [], "roles": roles or ["admin","hr"],
    })
    data["notifications"] = data["notifications"][:200]

EMP_FIELDS = ["id","full_name","email","phone","department","occupation","role","age","location",
              "employment_type","status","salary","salary_currency","contract_end",
              "start_date","date_added","manager_id","skills","emergency_contact"]

def export_csv_bytes(employees):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=EMP_FIELDS, extrasaction="ignore")
    w.writeheader()
    for eid, emp in employees.items():
        row = {"id": eid}
        row.update({k: emp.get(k, "") for k in EMP_FIELDS[1:]})
        if isinstance(row.get("skills"), list):
            row["skills"] = ", ".join(row["skills"])
        w.writerow(row)
    return buf.getvalue().encode()

def export_excel_bytes(employees):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Employees"
    hfill = PatternFill("solid", fgColor="07091a")
    hfont = Font(bold=True, color="C9A84C", size=11)
    headers = ["ID","Full Name","Email","Phone","Department","Occupation","Role","Age","Location",
               "Type","Status","Salary","Currency","Contract End","Start Date","Date Added",
               "Manager ID","Skills","Emergency Contact"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for ri, (eid, emp) in enumerate(employees.items(), 2):
        ws.cell(row=ri, column=1, value=eid)
        for ci, key in enumerate(EMP_FIELDS[1:], 2):
            val = emp.get(key, "")
            if isinstance(val, list): val = ", ".join(val)
            ws.cell(row=ri, column=ci, value=val)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 17
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def export_audit_csv_bytes(audit_log):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=["ts","user","action","target","detail"])
    w.writeheader()
    for entry in audit_log:
        w.writerow({k: entry.get(k,"") for k in ["ts","user","action","target","detail"]})
    return buf.getvalue().encode()
