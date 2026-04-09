"""NEXUS CYBER — Shared Data Layer v1.0
Cyber Security Workforce Management Platform
"""
import json, os, csv, io, hashlib, secrets
from datetime import datetime

DATA_FILE  = "nexus_data.json"
ROLE_LEVEL = {"admin": 4, "security_lead": 3, "team_lead": 2, "analyst": 1}
ROLE_LABEL = {"admin": "Security Administrator", "security_lead": "Security Lead", "team_lead": "Team Lead", "analyst": "Analyst"}

# ─── ROLE DEFINITIONS ────────────────────────────────────────────────────────
ROLE_DEFINITIONS = {
    "admin": {
        "label":       "Security Administrator",
        "icon":        "🛡️",
        "color":       "#00d4ff",
        "purpose":     "Full system control and security governance. Manages all users, security policies, and critical infrastructure access.",
        "responsibilities": [
            "Create, edit, and delete system user accounts",
            "Configure access roles and security clearance levels",
            "Approve or override any security decision",
            "Access all audit logs and security incident reports",
            "Manage team structure and organization-wide security policies",
        ],
        "level": 4,
    },
    "security_lead": {
        "label":       "Security Lead",
        "icon":        "🔐",
        "color":       "#00ff88",
        "purpose":     "Manages the complete security analyst lifecycle from onboarding to advanced clearance and all security operations workflows.",
        "responsibilities": [
            "Add, edit, and remove analyst records",
            "Approve or deny clearance upgrades and certification requests",
            "View and manage compensation and bonus data",
            "Run security team analytics and export reports",
            "Coordinate training, certifications, and compliance",
        ],
        "level": 3,
    },
    "team_lead": {
        "label":       "Team Lead",
        "icon":        "👁️",
        "color":       "#ffaa00",
        "purpose":     "Oversees team performance and day-to-day security operations within assigned teams.",
        "responsibilities": [
            "View all analyst records and team data",
            "Submit and track clearance upgrade requests for team members",
            "Submit certification requests on behalf of direct reports",
            "Monitor team availability, status, and workload",
            "Collaborate with Security Leads on hiring and performance reviews",
        ],
        "level": 2,
    },
    "analyst": {
        "label":       "Security Analyst",
        "icon":        "🔍",
        "color":       "#a855f7",
        "purpose":     "Standard analyst portal access for self-service tasks and personal record management.",
        "responsibilities": [
            "View own analyst profile and records",
            "Submit personal certification requests",
            "Request a clearance upgrade or role change",
            "View own compensation summary",
            "Update profile photo and change password",
        ],
        "level": 1,
    },
}

# ─── DEPARTMENT DEFINITIONS ───────────────────────────────────────────────────
DEPT_DEFINITIONS = {
    "Threat Intelligence": {
        "icon":        "🎯",
        "color":       "#ff4757",
        "purpose":     "Monitors and analyzes emerging cyber threats, threat actors, and attack vectors to provide proactive defense intelligence.",
        "responsibilities": [
            "Monitor dark web and threat actor forums",
            "Analyze malware samples and attack patterns",
            "Produce threat intelligence reports",
            "Maintain IOC databases and threat feeds",
            "Collaborate with industry threat-sharing groups",
        ],
        "kpis": ["Threat detection time", "IOC accuracy", "Intel report quality"],
    },
    "Security Operations (SOC)": {
        "icon":        "📡",
        "color":       "#00d4ff",
        "purpose":     "24/7 security monitoring and incident detection hub. First line of defense against cyber attacks.",
        "responsibilities": [
            "Monitor SIEM alerts and security events",
            "Triage and investigate security incidents",
            "Execute incident response playbooks",
            "Maintain security monitoring tools",
            "Coordinate with Incident Response team",
        ],
        "kpis": ["Mean time to detect (MTTD)", "Alert response time", "False positive rate"],
    },
    "Incident Response": {
        "icon":        "🚨",
        "color":       "#ff6b35",
        "purpose":     "Rapid response to security breaches and cyber attacks. Contains, eradicates, and recovers from incidents.",
        "responsibilities": [
            "Lead breach investigation and containment",
            "Perform digital forensics and evidence collection",
            "Coordinate with legal and compliance teams",
            "Develop and maintain IR playbooks",
            "Conduct post-incident reviews",
        ],
        "kpis": ["Mean time to respond (MTTR)", "Incident containment time", "Recovery success rate"],
    },
    "Penetration Testing": {
        "icon":        "🎭",
        "color":       "#a855f7",
        "purpose":     "Ethical hacking and vulnerability assessment to identify security weaknesses before attackers do.",
        "responsibilities": [
            "Perform network and web application pentests",
            "Conduct red team exercises",
            "Develop custom exploits and tools",
            "Report vulnerabilities with remediation guidance",
            "Validate security control effectiveness",
        ],
        "kpis": ["Vulnerabilities found", "Test coverage", "Remediation validation rate"],
    },
    "Compliance & Risk": {
        "icon":        "📋",
        "color":       "#22c55e",
        "purpose":     "Ensures organizational compliance with security standards, regulations, and risk management frameworks.",
        "responsibilities": [
            "Manage compliance audits (SOC2, ISO 27001, PCI-DSS)",
            "Conduct risk assessments and gap analysis",
            "Develop security policies and procedures",
            "Track regulatory requirements and changes",
            "Coordinate third-party security assessments",
        ],
        "kpis": ["Audit findings count", "Compliance score", "Policy adherence rate"],
    },
    "Security Engineering": {
        "icon":        "⚙️",
        "color":       "#3b82f6",
        "purpose":     "Designs, builds, and maintains security infrastructure, tools, and automation platforms.",
        "responsibilities": [
            "Build and maintain security tools and platforms",
            "Develop security automation and SOAR playbooks",
            "Manage SIEM, EDR, and security infrastructure",
            "Create detection rules and signatures",
            "Integrate security tools and APIs",
        ],
        "kpis": ["Tool uptime", "Automation coverage", "Detection rule accuracy"],
    },
}

def get_dept_definition(dept_name):
    if dept_name in DEPT_DEFINITIONS:
        return DEPT_DEFINITIONS[dept_name]
    return {
        "icon":  "🔒",
        "color": "#5a6080",
        "purpose": f"Responsible for {dept_name.lower()} security activities and operations.",
        "responsibilities": ["Department-specific security duties and responsibilities"],
        "kpis":  ["Security metrics", "Threat mitigation", "Team efficiency"],
    }

# ─── AUTH ─────────────────────────────────────────────────────────────────────
def hash_pw(pw):
    salt = secrets.token_hex(16)
    return f"{salt}:{hashlib.sha256((salt+pw).encode()).hexdigest()}"

def verify_pw(pw, stored):
    try:
        salt, h = stored.split(":", 1)
        return hashlib.sha256((salt+pw).encode()).hexdigest() == h
    except Exception:
        return False

# ─── DATA ─────────────────────────────────────────────────────────────────────
def _defaults():
    now = datetime.now().strftime("%Y-%m-%d")
    return {
        "meta": {"last_emp_num": 0},
        "system_users": {
            "admin":      {"password_hash": hash_pw("admin123"),  "role": "admin",          "full_name": "Security Administrator", "emp_id": None, "photo": None, "created": now},
            "sec_lead":   {"password_hash": hash_pw("lead123"),   "role": "security_lead",  "full_name": "Security Lead",          "emp_id": None, "photo": None, "created": now},
            "team_lead1": {"password_hash": hash_pw("tl123"),     "role": "team_lead",      "full_name": "Team Lead",              "emp_id": None, "photo": None, "created": now},
            "analyst1":   {"password_hash": hash_pw("analyst123"),"role": "analyst",        "full_name": "Security Analyst",       "emp_id": None, "photo": None, "created": now},
        },
        "employees": {},
        "departments": {
            "Threat Intelligence":      {"description": "Threat monitoring and intelligence analysis", "head_id": None},
            "Security Operations (SOC)":{"description": "24/7 security monitoring and incident detection", "head_id": None},
            "Incident Response":        {"description": "Breach response and digital forensics", "head_id": None},
            "Penetration Testing":      {"description": "Ethical hacking and vulnerability assessment", "head_id": None},
            "Compliance & Risk":        {"description": "Regulatory compliance and risk management", "head_id": None},
            "Security Engineering":     {"description": "Security tools and infrastructure engineering", "head_id": None},
        },
        "audit_log": [],
        "notifications": [],
        "leave_requests": [],
    }

def load_data():
    if not os.path.exists(DATA_FILE):
        d = _defaults(); save_data(d); return d
    try:
        with open(DATA_FILE) as f: return json.load(f)
    except Exception: return _defaults()

def save_data(data):
    with open(DATA_FILE, "w") as f: json.dump(data, f, indent=2)

def gen_emp_id(data):
    data["meta"]["last_emp_num"] = data["meta"].get("last_emp_num", 0) + 1
    return f"NXS-{datetime.now().year}-{data['meta']['last_emp_num']:04d}"

def audit(data, user, action, target, detail=""):
    data.setdefault("audit_log", []).insert(0, {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": user, "action": action, "target": target, "detail": detail,
    })
    data["audit_log"] = data["audit_log"][:1000]

def notify(data, ntype, message, roles=None):
    data.setdefault("notifications", []).insert(0, {
        "id": secrets.token_hex(8), "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": ntype, "message": message, "read_by": [], "roles": roles or ["admin","security_lead"],
    })
    data["notifications"] = data["notifications"][:200]

# Analyst fields (formerly employee fields)
EMP_FIELDS = ["id","full_name","email","phone","department","specialization","clearance_level","age","location",
              "employment_type","status","salary","salary_currency","contract_end",
              "start_date","date_added","team_lead_id","certifications","emergency_contact"]

def export_csv_bytes(employees):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=EMP_FIELDS, extrasaction="ignore")
    w.writeheader()
    for eid, emp in employees.items():
        row = {"id": eid}
        row.update({k: emp.get(k, "") for k in EMP_FIELDS[1:]})
        if isinstance(row.get("certifications"), list):
            row["certifications"] = ", ".join(row["certifications"])
        w.writerow(row)
    return buf.getvalue().encode()

def export_excel_bytes(employees):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Analysts"
    hfill = PatternFill("solid", fgColor="0a0f1a")
    hfont = Font(bold=True, color="00d4ff", size=11)
    headers = ["ID","Full Name","Email","Phone","Department","Specialization","Clearance Level","Age","Location",
               "Type","Status","Salary","Currency","Contract End","Start Date","Date Added",
               "Team Lead ID","Certifications","Emergency Contact"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for ri, (eid, emp) in enumerate(employees.items(), 2):
        ws.cell(row=ri, column=1, value=eid)
        for ci, key in enumerate(EMP_FIELDS[1:], 2):
            val = emp.get(key, "")
            if isinstance(val, list): val = ", ".join(val)
            ws.cell(row=ri, column=ci, value=val)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 17
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def export_audit_csv_bytes(audit_log):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=["ts","user","action","target","detail"])
    w.writeheader()
    for entry in audit_log:
        w.writerow({k: entry.get(k,"") for k in ["ts","user","action","target","detail"]})
    return buf.getvalue().encode()


def CLEARANCE_LEVELS():
    return None