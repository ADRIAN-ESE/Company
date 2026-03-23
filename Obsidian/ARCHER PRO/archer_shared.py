"""ARCHER ENTERPRISE — Shared Data Layer"""
import json, os, csv, io, hashlib, secrets
from datetime import datetime

DATA_FILE  = "archer_data.json"
ROLE_LEVEL = {"admin": 4, "hr": 3, "manager": 2, "staff": 1}
ROLE_LABEL = {"admin": "Admin", "hr": "HR Manager", "manager": "Manager", "staff": "Staff"}

def hash_pw(pw):
    salt = secrets.token_hex(16)
    return f"{salt}:{hashlib.sha256((salt+pw).encode()).hexdigest()}"

def verify_pw(pw, stored):
    try:
        salt, h = stored.split(":", 1)
        return hashlib.sha256((salt+pw).encode()).hexdigest() == h
    except Exception:
        return False

def _defaults():
    now = datetime.now().strftime("%Y-%m-%d")
    return {
        "meta": {"last_emp_num": 0},
        "system_users": {
            "admin":      {"password_hash": hash_pw("admin123"),  "role": "admin",   "full_name": "System Administrator", "emp_id": None, "photo": None, "created": now},
            "hr_manager": {"password_hash": hash_pw("hr123"),     "role": "hr",      "full_name": "HR Manager",           "emp_id": None, "photo": None, "created": now},
            "manager1":   {"password_hash": hash_pw("mgr123"),    "role": "manager", "full_name": "Department Manager",   "emp_id": None, "photo": None, "created": now},
            "staff1":     {"password_hash": hash_pw("staff123"),  "role": "staff",   "full_name": "Staff Member",         "emp_id": None, "photo": None, "created": now},
        },
        "employees": {},
        "departments": {
            "Administration":  {"description": "Administrative operations and corporate governance", "head_id": None},
            "Human Resources": {"description": "People operations, talent acquisition and development", "head_id": None},
            "Finance":         {"description": "Financial planning, accounting and treasury", "head_id": None},
            "Operations":      {"description": "Core business operations and service delivery", "head_id": None},
            "Technology":      {"description": "IT infrastructure, software and digital systems", "head_id": None},
        },
        "audit_log": [],
        "notifications": [],
        "leave_requests": [],
        "feedback": [],
    }

def _normalize(data: dict) -> dict:
    """Ensure critical fields are always the correct type, regardless of how data was saved."""
    # employees must be a dict {id: {...}} — fix if stored as a list
    emps = data.get("employees", {})
    if isinstance(emps, list):
        fixed = {}
        for emp in emps:
            if isinstance(emp, dict):
                eid = emp.get("id") or emp.get("emp_id")
                if not eid:
                    data["meta"]["last_emp_num"] = data["meta"].get("last_emp_num", 0) + 1
                    eid = f"ARCH-{datetime.now().year}-{data['meta']['last_emp_num']:04d}"
                emp.pop("id", None)
                emp.setdefault("promotion_requests", [])
                fixed[eid] = emp
        data["employees"] = fixed
    # ensure every employee value is a dict, not a list or primitive
    data["employees"] = {
        k: v for k, v in data.get("employees", {}).items()
        if isinstance(v, dict)
    }
    # leave_requests must be a list
    if not isinstance(data.get("leave_requests"), list):
        data["leave_requests"] = []
    # notifications must be a list
    if not isinstance(data.get("notifications"), list):
        data["notifications"] = []
    # audit_log must be a list
    if not isinstance(data.get("audit_log"), list):
        data["audit_log"] = []
    # feedback must be a list
    if not isinstance(data.get("feedback"), list):
        data["feedback"] = []
    return data

def load_data():
    if not os.path.exists(DATA_FILE):
        d = _defaults(); save_data(d); return d
    try:
        with open(DATA_FILE) as f:
            return _normalize(json.load(f))
    except Exception: return _defaults()

def save_data(data):
    with open(DATA_FILE, "w") as f: json.dump(data, f, indent=2)

def gen_emp_id(data):
    data["meta"]["last_emp_num"] = data["meta"].get("last_emp_num", 0) + 1
    return f"ARCH-{datetime.now().year}-{data['meta']['last_emp_num']:04d}"

def audit(data, user, action, target, detail=""):
    data.setdefault("audit_log", []).insert(0, {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": user, "action": action, "target": target, "detail": detail,
    })
    data["audit_log"] = data["audit_log"][:1000]

def notify(data, ntype, message, roles=None):
    data.setdefault("notifications", []).insert(0, {
        "id": secrets.token_hex(8), "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": ntype, "message": message, "read_by": [], "roles": roles or ["admin","hr"],
    })
    data["notifications"] = data["notifications"][:200]

EMP_FIELDS = ["id","full_name","email","department","occupation","role","purpose","age","location",
              "employment_type","status","salary","salary_currency","contract_end","date_added"]

def export_csv_bytes(employees):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=EMP_FIELDS, extrasaction="ignore")
    w.writeheader()
    for eid, emp in employees.items():
        row = {"id": eid}
        row.update({k: emp.get(k, "") for k in EMP_FIELDS[1:]})
        w.writerow(row)
    return buf.getvalue().encode()

def export_excel_bytes(employees):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Employees"
    hfill = PatternFill("solid", fgColor="07091a")
    hfont = Font(bold=True, color="C9A84C", size=11)
    headers = ["ID","Full Name","Email","Department","Occupation","Role","Age","Location",
               "Type","Status","Salary","Currency","Contract End","Date Added"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for ri, (eid, emp) in enumerate(employees.items(), 2):
        ws.cell(row=ri, column=1, value=eid)
        for ci, key in enumerate(EMP_FIELDS[1:], 2):
            ws.cell(row=ri, column=ci, value=emp.get(key, ""))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 17
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()
